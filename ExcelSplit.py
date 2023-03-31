import sys
from PySide6.QtWidgets import QApplication, QMainWindow, QFrame, QFileDialog, QMessageBox, QLabel, QLineEdit, QPushButton
from PySide6.QtGui import QFont, QValidator
import openpyxl.utils.cell as cell_utils
import traceback
from zipfile import BadZipFile
import openpyxl
import os
import copy
from openpyxl.utils import get_column_letter
from datetime import datetime
import pytz


def get_date_str():
    # 设置时区为北京时间
    tz = pytz.timezone('Asia/Shanghai')
    # 获取当前日期
    now = datetime.now(tz=tz)
    # 将日期格式化为"20230324"格式
    date_str = now.strftime('%Y%m%d')
    return date_str


def split_excel_by_column(input_file, start_row, end_row, key_column, out_put_path, date):

    # 获取工作簿和工作表对象
    wb = openpyxl.load_workbook(input_file)
    sheetname = wb.sheetnames[0]
    sheet = wb[sheetname]

    # 获取第2列第3到50行的值
    col_values = [sheet.cell(row=i, column=key_column).value for i in range(
        start_row, end_row+1)]
    # 创建一个字典，键为第2列的值，值为包含该值的所有行的列表
    data_dict = {}
    for i, value in enumerate(col_values):
        if value in data_dict:
            data_dict[value].append(i + start_row)
        else:
            data_dict[value] = [i + start_row]

    # 对于每个键值对，创建一个新的工作簿，将包含该值的行拷贝到新工作簿
    for key, rows in data_dict.items():
        if key == None:
            continue

        wb2 = openpyxl.Workbook()
        sheet2 = wb2.active
        sheet2.title = sheetname

        # Copy first two rows
        if start_row != 1:
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=start_row-1)):
                sheet2.row_dimensions[i +
                                      1].height = sheet.row_dimensions[i+1].height
                for j, cell in enumerate(row):
                    sheet2.column_dimensions[get_column_letter(
                        j+1)].width = sheet.column_dimensions[get_column_letter(j+1)].width
                    sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                    # 设置单元格格式
                    source_cell = sheet.cell(i+1, j+1)
                    target_cell = sheet2.cell(i+1, j+1)
                    if source_cell.has_style:
                        target_cell._style = copy.copy(source_cell._style)
                        target_cell.font = copy.copy(source_cell.font)
                        target_cell.border = copy.copy(source_cell.border)
                        target_cell.fill = copy.copy(source_cell.fill)
                        target_cell.number_format = copy.copy(
                            source_cell.number_format)
                        target_cell.protection = copy.copy(
                            source_cell.protection)
                        target_cell.alignment = copy.copy(
                            source_cell.alignment)

        # 拷贝包含该值的行
        for i, row in enumerate(rows):
            for j, cell in enumerate(sheet[row]):
                sheet2.column_dimensions[get_column_letter(
                    j + 1)].width = sheet.column_dimensions[get_column_letter(j + 1)].width
                sheet2.row_dimensions[i +
                                      start_row].height = sheet.row_dimensions[row].height
                sheet2.cell(row=i + start_row, column=j + 1, value=cell.value)

                # 设置单元格格式
                source_cell = sheet.cell(row=row, column=j+1)
                target_cell = sheet2.cell(row=i+start_row, column=j+1)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(
                        source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)

        # 跳过中间空白
        del_num_rows = end_row - start_row - len(rows) + 1

        # 处理合并的单元格
        for merged_cells in sheet.merged_cells.ranges:
            row_min, row_max, col_min, col_max = merged_cells.min_row, merged_cells.max_row, merged_cells.min_col, merged_cells.max_col
            if row_min > end_row:
                # 处理行列范围
                row_min -= del_num_rows
                row_max -= del_num_rows
            # 将行列范围转换为单元格对象，再获取单元格坐标
            cell_min = sheet2.cell(row=row_min, column=col_min)
            cell_max = sheet2.cell(row=row_max, column=col_max)
            cell_range = f'{cell_min.coordinate}:{cell_max.coordinate}'
            sheet2.merge_cells(cell_range)

        if sheet.max_row > end_row:
            # 拷贝剩余内容
            for i, row in enumerate(sheet.iter_rows(min_row=end_row+1, max_row=sheet.max_row)):
                sheet2.row_dimensions[i + end_row +
                                      1 - del_num_rows].height = sheet.row_dimensions[i+end_row+1].height
                for j, cell in enumerate(row):
                    sheet2.column_dimensions[get_column_letter(
                        j+1)].width = sheet.column_dimensions[get_column_letter(j+1)].width
                    sheet2.cell(row=i + end_row+1-del_num_rows, column=j +
                                1, value=cell.value)

                    # 设置单元格格式
                    source_cell = sheet.cell(i+end_row+1, j+1)
                    target_cell = sheet2.cell(i+end_row+1-del_num_rows, j+1)
                    if source_cell.has_style:
                        target_cell._style = copy.copy(source_cell._style)
                        target_cell.font = copy.copy(source_cell.font)
                        target_cell.border = copy.copy(source_cell.border)
                        target_cell.fill = copy.copy(source_cell.fill)
                        target_cell.number_format = copy.copy(
                            source_cell.number_format)
                        target_cell.protection = copy.copy(
                            source_cell.protection)
                        target_cell.alignment = copy.copy(
                            source_cell.alignment)

        # 删除除第一个sheet外的其他sheet
        if len(wb2.sheetnames) > 1:
            for s in wb2.sheetnames[1:]:
                del wb2[s]

        # 保存拷贝后的文档
        file_name = os.path.splitext(os.path.basename(input_file))[0]
        save_path = f"{out_put_path}/{key}{file_name}{date}.xlsx"
        wb2.save(save_path)
        wb2.close()

    wb.close()


class ExcelSplitor(QMainWindow):
    def __init__(self):
        super().__init__()
        # Set the default font for the application
        QApplication.setFont(QFont("微软雅黑", 10))
        self.setWindowTitle("表格分割")
        self.setGeometry(0, 0, 800, 600)
        self.move(int(QApplication.primaryScreen().size().width() / 2 - 400),
                  int(QApplication.primaryScreen().size().height() / 2 - 300))
        self.create_widgets()
        self.file_path = None
        self.start_rows = None
        self.end_rows = None
        self.split_col = None
        self.output_path = None

    def create_widgets(self):
        # Create a frame for the "Open File" button
        self.open_file_frame = QFrame(self)
        self.open_file_frame.setFrameStyle(QFrame.Panel | QFrame.Raised)
        self.open_file_frame.setGeometry(0, 0, 800, 600)

        # Create the "Open File" button
        self.open_file_button = QPushButton("打开文件", self.open_file_frame)
        self.open_file_button.setFont(QFont("微软雅黑", 20))
        self.open_file_button.setGeometry(0, 0, 800, 600)
        self.open_file_button.clicked.connect(self.open_file)
        self.open_file_frame.show()

    def open_file(self):
        self.file_path, _ = QFileDialog.getOpenFileName(
            self, "选择文件", "", "Excel files (*.xlsx;*.xls)")
        if self.file_path:
            self.start_rows = None
            self.end_rows = None
            self.split_col = None
            self.output_path = None
            self.open_file_frame.hide()
            self.show_input_boxes()

    def show_input_boxes(self):
        # Create a new frame to fill the window
        self.input_frame = QFrame(self)
        self.input_frame.setFrameStyle(QFrame.Panel | QFrame.Raised)
        self.input_frame.setGeometry(0, 0, 800, 600)

        self.split_column_label = QLabel(
            "请问根据哪一列的内容进行拆分(例如:1或A):", self.input_frame)
        self.split_column_label.setFont(QFont("微软雅黑", 20))
        self.split_column_label.setGeometry(0, 0, 800, 50)

        self.split_column_entry = QLineEdit(self.input_frame)
        self.split_column_entry.setFont(QFont("微软雅黑", 20))
        self.split_column_entry.setGeometry(0, 50, 800, 50)
        self.split_column_entry.setValidator(SplitColValidator())
        self.split_column_entry.textChanged.connect(
            self.on_split_column_entry_changed)

        # Create widgets inside the frame
        self.input_rows_label = QLabel(
            "请输入要拆分的起始行号和结束行号(1-1000):", self.input_frame)
        self.input_rows_label.setFont(QFont("微软雅黑", 20))
        self.input_rows_label.setGeometry(0, 100, 800, 50)

        self.start_row_label = QLabel(
            "起始行号: ", self.input_frame)
        self.start_row_label.setFont(QFont("微软雅黑", 20))
        self.start_row_label.setGeometry(10, 150, 150, 50)

        self.start_row_entry = QLineEdit(self.input_frame)
        self.start_row_entry.setFont(QFont("微软雅黑", 20))
        self.start_row_entry.setGeometry(160, 150, 230, 50)
        self.start_row_entry.setValidator(RowValidator())
        self.start_row_entry.textChanged.connect(
            self.on_start_row_entry_changed)

        self.end_row_label = QLabel(
            "结束行号: ", self.input_frame)
        self.end_row_label.setFont(QFont("微软雅黑", 20))
        self.end_row_label.setGeometry(410, 150, 150, 50)

        self.end_row_entry = QLineEdit(self.input_frame)
        self.end_row_entry.setFont(QFont("微软雅黑", 20))
        self.end_row_entry.setGeometry(560, 150, 230, 50)
        self.end_row_entry.setValidator(RowValidator())
        self.end_row_entry.textChanged.connect(self.on_end_row_entry_changed)

        self.output_path_button = QPushButton("选择输出文件夹", self.input_frame)
        self.output_path_button.setFont(QFont("微软雅黑", 20))
        self.output_path_button.setGeometry(0, 220, 800, 50)
        self.output_path_button.clicked.connect(self.choose_output_path)

        self.process_button = QPushButton("开始处理", self.input_frame)
        self.process_button.setFont(QFont("微软雅黑", 20))
        self.process_button.setGeometry(0, 290, 800, 50)
        self.process_button.clicked.connect(self.process_excel)

        # Set default values for input boxes
        self.start_row_entry.setText("2")
        self.end_row_entry.setText("50")
        self.split_column_entry.setText("A")

        # Show the input frame
        self.input_frame.show()

    def on_start_row_entry_changed(self, text):
        if text:
            self.start_rows = int(text)
        else:
            self.start_rows = None

    def on_end_row_entry_changed(self, text):
        if text:
            self.end_rows = int(text)
        else:
            self.end_rows = None

    def on_split_column_entry_changed(self, text):
        try:
            self.split_col = int(text)
        except ValueError:
            try:
                self.split_col = cell_utils.column_index_from_string(text)
            except ValueError:
                self.split_col = None

    def choose_output_path(self):
        self.output_path = QFileDialog.getExistingDirectory(self)
        if self.output_path:
            self.output_path_button.setText("已选择文件夹")

    def process_excel(self):
        if not all([self.file_path, self.start_rows, self.end_rows, self.split_col, self.output_path]):
            QMessageBox.critical(self, "处理异常", "请填写完整参数！")
            return

        try:
            # process
            if self.start_rows >= self.end_rows:
                raise RowException("起始行号大于或等于起始行号")
            date = get_date_str()
            split_excel_by_column(self.file_path, self.start_rows,
                                  self.end_rows, self.split_col, self.output_path, date)
            QMessageBox.information(self, "成功", "处理完成，请核验结果!")
        except BadZipFile:
            QMessageBox.warning(self, "提醒", "文件带有密码，请去除密码后再处理!")
        except RowException as e:
            QMessageBox.warning(self, "提醒", str(e))
            return
        except Exception as e:
            tb = traceback.format_exc()
            QMessageBox.critical(self, "错误", str(e) + "\n\n" + tb)

        self.split_column_label.hide()
        self.split_column_entry.hide()
        self.input_rows_label.hide()
        self.start_row_label.hide()
        self.start_row_entry.hide()
        self.end_row_label.hide()
        self.end_row_entry.hide()
        self.output_path_button.hide()
        self.process_button.hide()
        self.input_frame.hide()

        # Show the open file frame again
        self.open_file_frame.show()


class RowException(Exception):
    def __init__(self, message):
        self.message = message

    def __str__(self):
        return self.message


class RowValidator(QValidator):
    def validate(self, value, pos):
        if value.isdigit() and 1 <= int(value) <= 1000:
            return (QValidator.Acceptable, value, pos)
        elif value == "":
            return (QValidator.Acceptable, value, pos)
        else:
            return (QValidator.Invalid, value, pos)


class SplitColValidator(QValidator):
    def validate(self, value, pos):
        try:
            if value.isdigit():
                if 1 <= int(value) <= 100:
                    return (QValidator.Acceptable, value, pos)
                else:
                    return (QValidator.Invalid, value, pos)
            elif value.isalpha():
                try:
                    cell_utils.column_index_from_string(value)
                    return (QValidator.Acceptable, value, pos)
                except:
                    return (QValidator.Invalid, value, pos)
            elif value == "":
                return (QValidator.Acceptable, value, pos)
            else:
                return (QValidator.Invalid, value, pos)
        except:
            return (QValidator.Invalid, value, pos)


if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = ExcelSplitor()
    window.show()
    sys.exit(app.exec())
